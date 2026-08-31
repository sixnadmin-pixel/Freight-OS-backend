"""
Parse an RFQ tender spreadsheet into a list of routes.

Targets the 'Guess Tender' layout:
  - a header row containing a UNLOCODE column
  - one data row per origin port
  - the destination (POD) in a merged header cell above the rate columns
"""
from io import BytesIO
import openpyxl

import json
from openai import OpenAI
from modules.authen import config


def _find_header_row(ws, max_scan: int = 15):
    """Return (row_index, {lowercased header: column_index})."""
    for r in range(1, min(ws.max_row, max_scan) + 1):
        headers = {}
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip():
                headers[v.strip().lower()] = c
        if any("unlocode" in h for h in headers):
            return r, headers
    return None, {}


def _find_pod(ws) -> str | None:
    """The destination lives in a merged header cell. 'POD' alone means unfilled."""
    for m in ws.merged_cells.ranges:
        v = ws.cell(m.min_row, m.min_col).value
        if isinstance(v, str):
            v = v.strip()
            if v and v.upper() != "POD":
                return v
    return None


def sheet_to_text(file_bytes: bytes, max_rows: int = 250, max_cols: int = 15) -> str:
    """Flatten every sheet into a compact text grid the LLM can read."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    out = []
    for ws in wb.worksheets:
        out.append(f"### SHEET: {ws.title}  ({ws.max_row} rows x {ws.max_column} cols)")
        merged = [str(m) for m in ws.merged_cells.ranges]
        if merged:
            out.append(f"merged cells: {', '.join(merged[:12])}")
        for r in range(1, min(ws.max_row, max_rows) + 1):
            cells = []
            for c in range(1, min(ws.max_column, max_cols) + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    cells.append(f"{chr(64 + c)}{r}={str(v).strip()[:40]}")
            if cells:
                out.append(" | ".join(cells))
        if ws.max_row > max_rows:
            out.append(f"... ({ws.max_row - max_rows} more rows, same shape)")
    return "\n".join(out)




_PROMPT = """You are reading a freight forwarding RFQ / tender spreadsheet.
CONTEXT: This is for CL Synergy, a freight forwarder based in Colombo, Sri Lanka
(UN/LOCODE LKCMB). Most of their business is EXPORT out of Colombo. So when a
sheet lists many foreign ports and has no origin column, those ports are almost
certainly DESTINATIONS, with Colombo as the unstated origin. Only treat a list
as origins when the sheet says so explicitly — a column headed "Port of Loading",
"POL", "Origin", or a UNLOCODE column of loading ports.

The customer wants shipping prices for a set of routes. Each route has an ORIGIN
port and a DESTINATION port. Usually one end is fixed for the whole sheet (often
in a merged header cell) and the other varies per row.

Return ONLY JSON, no prose:
{
  "fixed_origin": "<UN/LOCODE or port name, or null>",
  "fixed_destination": "<UN/LOCODE or port name, or null>",
  "rows": [
    {"row": <sheet row number>,
     "origin": "<UN/LOCODE if the sheet gives one, else the port name>",
     "destination": "<same, or null if it's the fixed end>",
     "country": "<country if shown, else null>"}
  ],
  "notes": "<anything unclear, one short sentence>"
}

Rules:
- Only include rows that are actual routes. Skip headers, totals, blank rows.
- If the sheet contains a column of UN/LOCODEs (5-character codes like CNSGH,
  LKCMB, DEHAM), you MUST return those codes verbatim. Prefer them over any
  human-readable port name in another column.
- Only return a port NAME when the sheet gives no code for that row.
- Never invent a UN/LOCODE.
- If it only gives a name, correct obvious misspellings to the real port name
  (e.g. "Huston" -> "Houston", "Cohin" -> "Cochin", "Nava Sheva" -> "Nhava Sheva").
- Do NOT invent UN/LOCODEs. If you only have a name, return the corrected NAME.
- Strip prefixes like "CIF -" from port names.
- If the sheet has NO origin column, return origin: null for every row.
  NEVER use a Country or Region column as the origin — those are attributes
  of the port, not a separate port.
- Likewise if there is no destination column, return destination: null.
- The fixed end (the one not in the rows) may be missing entirely. That is
  fine — return null and the user will supply it.

"""


def parse_with_llm(file_bytes: bytes) -> dict:
    if not config.OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API is not set")

    client = OpenAI(api_key=config.OPENAI_API_KEY)
    resp = client.chat.completions.create(
        model=config.OPENAI_MODEL,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": _PROMPT},
            {"role": "user", "content": sheet_to_text(file_bytes)},
        ],
    )
    data = json.loads(resp.choices[0].message.content)

    fixed_o = (data.get("fixed_origin") or "").strip() or None
    fixed_d = (data.get("fixed_destination") or "").strip() or None

    rows = []
    for r in data.get("rows", []):
        o = (r.get("origin") or fixed_o or "").strip()
        d = (r.get("destination") or fixed_d or "").strip()
        if not o and not d:
            continue
        rows.append({
            "row": r.get("row") or 0,
            "origin": o,
            "destination": d,
            "country": (r.get("country") or "").strip(),
        })

    return {"origin": fixed_o, "destination": fixed_d, "rows": rows,
            "skipped": [], "notes": data.get("notes", "")}




def parse_tender_keyword(file_bytes: bytes) -> dict:
    """
    Returns:
      { "destination": str | None,
        "rows": [ {"row": int, "unlocode": str, "port_name": str, "country": str} ],
        "skipped": [ {"row": int, "reason": str} ] }
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    ws = None
    for sheet in wb.worksheets:
        r, _ = _find_header_row(sheet)
        if r:
            ws = sheet
            break
    if ws is None:
        raise ValueError("No sheet with a UNLOCODE column was found")

    header_row, headers = _find_header_row(ws)

    def col(*names):
        for n in names:
            for h, idx in headers.items():
                if n in h:
                    return idx
        return None

    c_code = col("unlocode")
    c_name = col("port of loading", "port")
    c_ctry = col("country")

    rows, skipped = [], []
    for r in range(header_row + 1, ws.max_row + 1):
        code = ws.cell(r, c_code).value if c_code else None
        if code is None or not str(code).strip():
            continue
        code = str(code).strip().upper()
        if len(code) != 5:
            skipped.append({"row": r, "reason": f"'{code}' is not a 5-character UN/LOCODE"})
            continue
        rows.append({
            "row": r,
            "origin": code,          # this layout varies the ORIGIN per row
            "destination": "",       # fixed for the sheet — see below
            "country": str(ws.cell(r, c_ctry).value or "").strip() if c_ctry else "",
        })

    pod = _find_pod(ws)
    for r in rows:
        r["destination"] = pod or ""
    return {"origin": None, "destination": pod, "rows": rows, "skipped": skipped}



def parse_tender_excel(file_bytes: bytes) -> dict:
    """
    Deterministic parser first — when the sheet has a UNLOCODE column it gives
    exact codes and costs nothing. Fall back to the LLM for any other layout.
    """
    try:
        result = parse_tender_keyword(file_bytes)
        if result.get("rows"):
            return result
    except Exception as e:
        print(f"[excel_processor] keyword parse failed ({e}), trying LLM")

    return parse_with_llm(file_bytes)

